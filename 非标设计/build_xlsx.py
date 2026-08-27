#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成《非标设计》选型表：伺服步进丝杠齿条选型.xlsx。公式在表内，用 Excel / WPS 打开即算。"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent / "伺服步进丝杠齿条选型.xlsx"

YELLOW = PatternFill("solid", fgColor="FFF3CD")
GREEN = PatternFill("solid", fgColor="D1E7DD")
GRAY = PatternFill("solid", fgColor="F1F3F5")
BLUE = PatternFill("solid", fgColor="D6EAF8")
RED_FONT = Font(color="C0392B", name="微软雅黑", size=11)
GREEN_FONT = Font(color="196F3D", name="微软雅黑", size=11)
TITLE = Font(name="微软雅黑", size=16, bold=True, color="1B4F3A")
H = Font(name="微软雅黑", size=12, bold=True)
N = Font(name="微软雅黑", size=11)
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def sty(cell, font=N, fill=None, unlock=False, wrap=False, fmt=None):
    cell.font = font
    cell.alignment = Alignment(wrap_text=wrap, vertical="center")
    cell.border = THIN
    if fill:
        cell.fill = fill
    cell.protection = Protection(locked=not unlock)
    if fmt:
        cell.number_format = fmt


def colw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    wb = Workbook()

    # ---- 目录丝杠（先建，推荐用 MATCH 第一档可用）----
    wsS = wb.active
    wsS.title = "目录丝杠"
    heads = ["型号", "直径mm", "导程mm", "动载kN", "现货", "转速", "转矩", "丝杠长", "临界转速", "可用", "BK", "BF", "座内角接触", "座内深沟球"]
    for i, h in enumerate(heads, 1):
        c = wsS.cell(1, i, h)
        sty(c, H, BLUE)
    screws = [
        # 现货优先、常用在前
        ("SFU2510-3", 25, 10, 17.8, 1, 20, 20, "7004AC×2", "6004ZZ"),
        ("SFU4010-3", 40, 10, 37.4, 1, 25, 25, "7205AC×2", "6005ZZ"),
        ("SFU2010-3", 20, 10, 11.0, 1, 15, 15, "7002AC×2", "6002ZZ"),
        ("SFU2505-3", 25, 5, 17.8, 1, 20, 20, "7004AC×2", "6004ZZ"),
        ("SFU4005-3", 40, 5, 37.4, 1, 25, 25, "7205AC×2", "6005ZZ"),
        ("SFU2005-3", 20, 5, 11.0, 1, 15, 15, "7002AC×2", "6002ZZ"),
        ("SFU5010-3", 50, 10, 54.0, 1, 30, 30, "7206AC×2", "6006ZZ"),
        ("SFU3205-3", 32, 5, 27.0, 1, 20, 20, "7004AC×2", "6004ZZ"),
        ("SFU1605-3", 16, 5, 7.8, 1, 12, 10, "7001AC×2", "6001ZZ"),
        ("SFU3210-3", 32, 10, 27.0, 0, 20, 20, "7004AC×2", "6004ZZ"),
        ("SFU3220-3", 32, 20, 27.0, 0, 20, 20, "7004AC×2", "6004ZZ"),
        ("SFU4020-3", 40, 20, 37.4, 0, 25, 25, "7205AC×2", "6005ZZ"),
    ]
    # 转速=输入速度*1000/导程 ；转矩=工作推力*导程/1000/(2*PI*效率)
    for r, row in enumerate(screws, 2):
        model, d, lead, ca, stock, bkj, bfj, ang, gro = row
        wsS.cell(r, 1, model)
        wsS.cell(r, 2, d)
        wsS.cell(r, 3, lead)
        wsS.cell(r, 4, ca)
        wsS.cell(r, 5, stock)
        wsS.cell(r, 6, f"=输入!$B$9*1000/C{r}")
        wsS.cell(r, 7, f"=计算!$B$6*C{r}/1000/(2*PI()*输入!$B$15)")
        wsS.cell(r, 8, f"=输入!$B$10+400")
        wsS.cell(r, 9, f"=15.1E8*B{r}/H{r}^2*0.8")
        # 可用：现货 + 转速<3000 + 转速<临界 + 行程长轴径>=25 + 导程5且转速>1400否 + 推力下轴径
        wsS.cell(
            r,
            10,
            f'=IF(AND(E{r}=1,F{r}<=3000,F{r}<I{r},'
            f'IF(输入!$B$10>1800,B{r}>=25,TRUE),'
            f'IF(AND(C{r}=5,F{r}>1400),FALSE,TRUE),'
            f'IF(计算!$B$6>800,B{r}>=25,TRUE)),1,0)',
        )
        wsS.cell(r, 11, f'=IF(B{r}<=16,"BK12",IF(B{r}<=20,"BK15",IF(B{r}<=32,"BK20",IF(B{r}<=40,"BK25","BK30"))))')
        wsS.cell(r, 12, f'=IF(B{r}<=16,"BF12",IF(B{r}<=20,"BF15",IF(B{r}<=32,"BF20",IF(B{r}<=40,"BF25","BF30"))))')
        wsS.cell(r, 13, ang)
        wsS.cell(r, 14, gro)
        for c in range(1, 15):
            sty(wsS.cell(r, c), fill=GRAY, fmt="0.00" if c in (4, 6, 7, 8, 9) else "General")
    wsS.cell(20, 1, "手选下拉用")
    sty(wsS.cell(20, 1), H)
    wsS.cell(21, 1, "（用推荐）")
    for r, row in enumerate(screws, 22):
        wsS.cell(r, 1, row[0])
    colw(wsS, [16, 12, 12, 12, 8, 12, 12, 12, 14, 8, 10, 10, 16, 14])
    wsS.sheet_state = "visible"

    # ---- 目录电机 ----
    wsM = wb.create_sheet("目录电机")
    mh = ["类型", "名称", "机座", "额定或静矩", "峰值或持矩", "额定转速", "功率档kW", "丝杠方案可用矩", "丝杠运行SF", "丝杠启动SF", "丝杠可用", "齿条方案可用矩", "齿条运行SF", "齿条启动SF", "齿条可用"]
    for i, h in enumerate(mh, 1):
        sty(wsM.cell(1, i, h), H, BLUE)
    motors = [
        ("伺服", "0.4 kW 伺服带刹车", "60", 1.27, 3.81, 3000, 0.4),
        ("伺服", "0.75 kW 伺服带刹车", "80", 2.39, 7.16, 3000, 0.75),
        ("伺服", "1.0 kW 伺服带刹车", "80", 3.18, 9.54, 3000, 1.0),
        ("伺服", "1.5 kW 伺服带刹车", "100", 4.77, 14.3, 3000, 1.5),
        ("伺服", "2.0 kW 伺服带刹车", "100", 6.37, 19.1, 3000, 2.0),
        ("伺服", "3.0 kW 伺服带刹车", "130", 9.55, 28.6, 3000, 3.0),
        ("步进", "57 闭环带刹车", "57", 2.2, 2.2, 1500, 0.2),
        ("步进", "86 闭环 8 N·m 带刹车", "86", 8, 8, 1500, 0.55),
        ("步进", "86 闭环 12 N·m 带刹车", "86", 12, 12, 1500, 0.75),
        ("步进", "110 闭环 20 N·m 带刹车", "110", 20, 20, 1500, 1.2),
        ("步进", "130 闭环 25 N·m 带刹车", "130", 25, 25, 1500, 1.8),
    ]
    def t_step(cell_hold, n_ref):
        # 闭环步进随转速掉矩（经验曲线）
        return f"{cell_hold}/(1+({n_ref}/600)^1.6)"

    for r, m in enumerate(motors, 2):
        for i, v in enumerate(m, 1):
            wsM.cell(r, i, v)
        # 丝杠方案 n=计算!B13 T=计算!B14 Tstart=计算!B7对应转矩
        nS, tS, t0S = "计算!$B$13", "计算!$B$14", "计算!$B$15"
        nR, tR, t0R = "计算!$B$20", "计算!$B$21", "计算!$B$22"
        if m[0] == "伺服":
            wsM.cell(r, 8, f'=IF({nS}<=F{r},D{r},D{r}*F{r}/MAX({nS},1))')
            wsM.cell(r, 12, f'=IF({nR}<=F{r},D{r},D{r}*F{r}/MAX({nR},1))')
        else:
            wsM.cell(r, 8, f"={t_step(f'D{r}', nS)}")
            wsM.cell(r, 12, f"={t_step(f'D{r}', nR)}")
        wsM.cell(r, 9, f"=H{r}/MAX({tS},0.01)")
        wsM.cell(r, 10, f"=E{r}/MAX({t0S},0.01)")
        wsM.cell(r, 11, f"=IF(AND(I{r}>=输入!$B$13,J{r}>=1.5,{nS}<=F{r}*1.05),1,0)")
        wsM.cell(r, 13, f"=L{r}/MAX({tR},0.01)")
        wsM.cell(r, 14, f"=E{r}/MAX({t0R},0.01)")
        wsM.cell(r, 15, f"=IF(AND(M{r}>=输入!$B$13,N{r}>=1.5,{nR}<=F{r}*1.05),1,0)")
        for c in range(1, 16):
            sty(wsM.cell(r, c), fill=GRAY, fmt="0.00" if c >= 4 else "General")
    wsM.cell(20, 1, "伺服下拉")
    wsM.cell(21, 1, "（用推荐）")
    for r, m in enumerate(motors, 22):
        if m[0] == "伺服":
            wsM.cell(r, 1, m[1])
    wsM.cell(20, 2, "步进下拉")
    wsM.cell(21, 2, "（用推荐）")
    i = 22
    for m in motors:
        if m[0] == "步进":
            wsM.cell(i, 2, m[1])
            i += 1
    colw(wsM, [10, 28, 10, 14, 14, 12, 12, 16, 12, 12, 10, 16, 12, 12, 10])

    # ---- 输入 ----
    wsI = wb.create_sheet("输入", 0)
    wsI["A1"] = "非标设计　伺服 / 步进 / 丝杠 / 齿轮齿条选型"
    sty(wsI["A1"], TITLE)
    wsI.merge_cells("A1:D1")
    wsI["A2"] = "黄格可改。改完先看「推荐」：结果旁边就是公式和代入数字。完整 22 步在「计算过程」。架子重量填实重，不要用别人估的 640 kg。"
    sty(wsI["A2"], N, wrap=True)
    wsI.merge_cells("A2:D2")
    wsI.row_dimensions[2].height = 36

    labels = [
        (4, "安装方向", "水平", "水平 / 垂直向上 / 垂直向下 / 倾斜上坡 / 倾斜下坡"),
        (5, "倾角_度", 0, "仅倾斜时填写，水平垂直填 0"),
        (6, "工件质量_kg", 151, "例如 3 根 Φ219 管约 151"),
        (7, "机构质量_kg", 200, "架子+滑座，按你的设计填"),
        (8, "额外阻力_N", 0, "刮屑、密封等，没有填 0"),
        (9, "速度_m_min", 8, "任务书上限 15，上料常用 6～8"),
        (10, "行程_mm", 2500, "影响丝杠会不会垂"),
        (11, "加速时间_s", 0.4, "越小加速力越大"),
        (12, "摩擦系数", 0.1, "铜套约 0.1，直线导轨约 0.02"),
        (13, "目标安全系数", 2, "运行力矩相对需要力矩"),
        (14, "传动计算", "两种都算", "自动 / 只要丝杠 / 只要齿条 / 两种都算"),
        (15, "丝杠效率", 0.90, "滚珠丝杠常用 0.90"),
        (16, "齿条型式", "斜齿", "斜齿或直齿"),
        (17, "齿条效率", 0.94, "斜齿 0.94，直齿改 0.90"),
        (18, "齿条模数", 2, "给定清单：1.5 / 2 / 2.5 / 3"),
        (19, "齿条齿数", 25, "给定清单：20 / 25 / 30"),
        (20, "齿条减速比", 10, "1=直连（力矩大，目录常无满足项）。常用 8～15"),
        (21, "减速机效率", 0.94, "直连（减速比=1）时公式按 1 计；行星常用 0.94"),
    ]
    wsI["A3"] = "项目"
    wsI["B3"] = "数值"
    wsI["C3"] = "说明"
    for col in range(1, 4):
        sty(wsI.cell(3, col), H, BLUE)
    for row, lab, val, note in labels:
        wsI.cell(row, 1, lab)
        wsI.cell(row, 2, val)
        wsI.cell(row, 3, note)
        sty(wsI.cell(row, 1), H, GRAY)
        sty(wsI.cell(row, 2), N, YELLOW, unlock=True)
        sty(wsI.cell(row, 3), Font(name="微软雅黑", size=10, color="666666"))
        if row in (5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 17, 18, 19, 20, 21):
            wsI.cell(row, 2).number_format = "0.00"

    dv_dir = DataValidation(type="list", formula1='"水平,垂直向上,垂直向下,倾斜上坡,倾斜下坡"', allow_blank=False)
    dv_dir.add("B4")
    wsI.add_data_validation(dv_dir)
    dv_tr = DataValidation(type="list", formula1='"自动,只要丝杠,只要齿条,两种都算"', allow_blank=False)
    dv_tr.add("B14")
    wsI.add_data_validation(dv_tr)
    dv_rack = DataValidation(type="list", formula1='"斜齿,直齿"', allow_blank=False)
    dv_rack.add("B16")
    wsI.add_data_validation(dv_rack)
    dv_m = DataValidation(type="list", formula1='"1.5,2,2.5,3"', allow_blank=False)
    dv_m.add("B18")
    wsI.add_data_validation(dv_m)
    dv_z = DataValidation(type="list", formula1='"20,25,30"', allow_blank=False)
    dv_z.add("B19")
    wsI.add_data_validation(dv_z)
    dv_i = DataValidation(type="list", formula1='"1,5,8,10,12,15,20"', allow_blank=False)
    dv_i.add("B20")
    wsI.add_data_validation(dv_i)

    wsI["A23"] = "手选（可改下拉；选「用推荐」则跟推荐走）"
    sty(wsI["A23"], H)
    wsI.merge_cells("A23:C23")
    hand = [
        (24, "手选丝杠", "（用推荐）"),
        (25, "手选伺服", "（用推荐）"),
        (26, "手选步进", "（用推荐）"),
        (27, "手选齿条模数", "（用推荐）"),
        (28, "手选齿条齿数", "（用推荐）"),
        (29, "手选齿条减速比", "（用推荐）"),
    ]
    for row, lab, val in hand:
        wsI.cell(row, 1, lab)
        wsI.cell(row, 2, val)
        sty(wsI.cell(row, 1), H, GRAY)
        sty(wsI.cell(row, 2), N, YELLOW, unlock=True)
    dv_s = DataValidation(type="list", formula1="=目录丝杠!$A$21:$A$33", allow_blank=False)
    dv_s.add("B24")
    wsI.add_data_validation(dv_s)
    dv_sv = DataValidation(type="list", formula1="=目录电机!$A$21:$A$27", allow_blank=False)
    dv_sv.add("B25")
    wsI.add_data_validation(dv_sv)
    dv_st = DataValidation(type="list", formula1="=目录电机!$B$21:$B$26", allow_blank=False)
    dv_st.add("B26")
    wsI.add_data_validation(dv_st)
    dv_hm = DataValidation(type="list", formula1='"（用推荐）,1.5,2,2.5,3"', allow_blank=False)
    dv_hm.add("B27")
    wsI.add_data_validation(dv_hm)
    dv_hz = DataValidation(type="list", formula1='"（用推荐）,20,25,30"', allow_blank=False)
    dv_hz.add("B28")
    wsI.add_data_validation(dv_hz)
    dv_hi = DataValidation(type="list", formula1='"（用推荐）,1,5,8,10,12,15,20"', allow_blank=False)
    dv_hi.add("B29")
    wsI.add_data_validation(dv_hi)

    colw(wsI, [22, 28, 55])
    wsI.row_dimensions[1].height = 24
    wsI.freeze_panes = "A4"

    # ---- 计算 ----
    wsC = wb.create_sheet("计算")
    wsC["A1"] = "中间计算（可看公式，不必改）"
    sty(wsC["A1"], H)
    calcs = [
        (2, "总质量kg", "=输入!B6+输入!B7"),
        (3, "速度m_s", "=输入!B9/60"),
        (4, "加速度", "=B3/MAX(输入!B11,0.05)"),
        (5, "等效倾角度", '=IF(输入!B4="水平",0,IF(输入!B4="垂直向上",90,IF(输入!B4="垂直向下",-90,IF(输入!B4="倾斜上坡",输入!B5,-输入!B5))))'),
        (6, "工作推力N", "=ABS(B2*9.81*SIN(B5*PI()/180)+输入!B12*B2*9.81*COS(B5*PI()/180)+B2*B4+输入!B8)"),
        (7, "沿程重力N", "=B2*9.81*SIN(B5*PI()/180)"),
        (8, "摩擦N", "=输入!B12*B2*9.81*ABS(COS(B5*PI()/180))"),
        (9, "加速力N", "=B2*B4"),
        (10, "是否倒拖", '=IF(B2*9.81*SIN(B5*PI()/180)+B2*B4+输入!B8<0,"是","否")'),
        (11, "推荐丝杠行", "=IFERROR(MATCH(1,目录丝杠!J2:J13,0),1)"),
        (12, "推荐丝杠", "=INDEX(目录丝杠!A2:A13,B11)"),
        (13, "丝杠转速", "=INDEX(目录丝杠!F2:F13,B11)"),
        (14, "丝杠工作转矩", "=INDEX(目录丝杠!G2:G13,B11)"),
        (15, "丝杠启动转矩", "=B6*INDEX(目录丝杠!C2:C13,B11)/1000/(2*PI()*输入!B15)*IF(输入!B4=\"水平\",MAX(输入!B12,0.15)/MAX(输入!B12,0.01),1)"),
        (16, "推荐BK", "=INDEX(目录丝杠!K2:K13,B11)"),
        (17, "推荐BF", "=INDEX(目录丝杠!L2:L13,B11)"),
        (18, "角接触", "=INDEX(目录丝杠!M2:M13,B11)"),
        (19, "深沟球", "=INDEX(目录丝杠!N2:N13,B11)"),
        (20, "齿条电机转速", "=输入!B9*1000/(PI()*输入!B18*输入!B19)*MAX(输入!B20,1)"),
        (21, "齿条电机转矩", "=B6*(输入!B18*输入!B19)/2000/输入!B17/MAX(输入!B20,1)/IF(输入!B20<=1,1,输入!B21)"),
        (22, "齿条启动转矩", "=B15/MAX(B14,0.01)*B21"),
        (23, "有效齿条模数", '=IF(输入!B27="（用推荐）",输入!B18,VALUE(输入!B27))'),
        (24, "有效齿条齿数", '=IF(输入!B28="（用推荐）",输入!B19,VALUE(输入!B28))'),
        (38, "有效减速比", '=IF(输入!B29="（用推荐）",输入!B20,VALUE(输入!B29))'),
        (25, "手选后齿条转速", "=输入!B9*1000/(PI()*B23*B24)*MAX(B38,1)"),
        (26, "手选后齿条转矩", "=B6*(B23*B24)/2000/输入!B17/MAX(B38,1)/IF(B38<=1,1,输入!B21)"),
        (27, "有效丝杠", '=IF(输入!B24="（用推荐）",B12,输入!B24)'),
        (28, "有效丝杠转速", '=IFERROR(INDEX(目录丝杠!F2:F13,MATCH(B27,目录丝杠!A2:A13,0)),B13)'),
        (29, "有效丝杠转矩", '=IFERROR(INDEX(目录丝杠!G2:G13,MATCH(B27,目录丝杠!A2:A13,0)),B14)'),
        (30, "推荐伺服", '=IFERROR(INDEX(目录电机!B2:B7,MATCH(1,目录电机!K2:K7,0)),"无满足项，请降速或加大电机")'),
        (31, "推荐步进", '=IFERROR(INDEX(目录电机!B8:B12,MATCH(1,目录电机!K8:K12,0)),"无满足项，请降速或改110/130")'),
        (32, "齿条推荐伺服", '=IFERROR(INDEX(目录电机!B2:B7,MATCH(1,目录电机!O2:O7,0)),"无满足项，请加大减速比")'),
        (33, "齿条推荐步进", '=IFERROR(INDEX(目录电机!B8:B12,MATCH(1,目录电机!O8:O12,0)),"无满足项，请加大减速比")'),
        (34, "有效伺服", '=IF(输入!B25="（用推荐）",B30,输入!B25)'),
        (35, "有效步进", '=IF(输入!B26="（用推荐）",B31,输入!B26)'),
        (36, "机械功率kW", "=B14*B13/9550"),
        (37, "齿条功率kW", "=B21*B20/9550"),
        (39, "齿轮轴转速", "=输入!B9*1000/(PI()*输入!B18*输入!B19)"),
        (40, "推荐导程mm", "=INDEX(目录丝杠!C2:C13,B11)"),
        (41, "sinθ", "=SIN(B5*PI()/180)"),
        (42, "cosθ", "=COS(B5*PI()/180)"),
        (43, "2πη", "=2*PI()*输入!B15"),
        (44, "齿条节圆半径mm", "=输入!B18*输入!B19/2"),
        (45, "推力未取绝对值N", "=B7+B8+B9+输入!B8"),
    ]
    wsC.cell(1, 1, "符号")
    wsC.cell(1, 2, "数值")
    sty(wsC.cell(1, 1), H, BLUE)
    sty(wsC.cell(1, 2), H, BLUE)
    for row, name, formula in calcs:
        wsC.cell(row, 1, name)
        wsC.cell(row, 2, formula)
        sty(wsC.cell(row, 1), N, GRAY)
        sty(wsC.cell(row, 2), N, GRAY, fmt="0.00")
    colw(wsC, [22, 55])

    # ---- 推荐（结果 + 公式 + 代入，同一页）----
    wsR = wb.create_sheet("推荐", 1)
    wsR["A1"] = "非标设计　推荐结果（只读；右边是公式和代入）"
    sty(wsR["A1"], TITLE)
    wsR.merge_cells("A1:D1")
    wsR["A2"] = '=输入!B4&"　总质量 "&TEXT(计算!B2,"0.00")&" kg　速度 "&TEXT(输入!B9,"0.00")&" m/min　工作推力 "&TEXT(计算!B6,"0.0")&" N。完整 22 步见「计算过程」。'
    sty(wsR["A2"], N, wrap=True)
    wsR.merge_cells("A2:D2")
    wsR.row_dimensions[2].height = 28
    rec = [
        (4, "总质量 kg", "=计算!B2",
         "m = 工件 + 机构",
         '=TEXT(输入!B6,"0.00")&" + "&TEXT(输入!B7,"0.00")&" = "&TEXT(计算!B2,"0.00")&" kg"'),
        (5, "速度 m/s", "=计算!B3",
         "v = 速度(m/min) ÷ 60",
         '=TEXT(输入!B9,"0.00")&" / 60 = "&TEXT(计算!B3,"0.000")&" m/s"'),
        (6, "加速度 m/s²", "=计算!B4",
         "a = v ÷ 加速时间",
         '=TEXT(计算!B3,"0.000")&" / "&TEXT(输入!B11,"0.00")&" = "&TEXT(计算!B4,"0.000")&" m/s²"'),
        (7, "等效倾角 °", "=计算!B5",
         "水平0；垂直上+90；垂直下−90；倾斜取填角（下坡为负）",
         '=输入!B4&" → "&TEXT(计算!B5,"0.0")&" °　sin="&TEXT(计算!B41,"0.000")&"　cos="&TEXT(计算!B42,"0.000")'),
        (8, "沿程重力 N", "=计算!B7",
         "Fg = m g sinθ",
         '=TEXT(计算!B2,"0.00")&" × 9.81 × SIN("&TEXT(计算!B5,"0.0")&"°) = "&TEXT(计算!B7,"0.0")&" N"'),
        (9, "摩擦力 N", "=计算!B8",
         "Ff = μ m g |cosθ|",
         '=TEXT(输入!B12,"0.00")&" × "&TEXT(计算!B2,"0.00")&" × 9.81 × |"&TEXT(计算!B42,"0.000")&"| = "&TEXT(计算!B8,"0.0")&" N"'),
        (10, "加速力 N", "=计算!B9",
         "Fa = m a",
         '=TEXT(计算!B2,"0.00")&" × "&TEXT(计算!B4,"0.000")&" = "&TEXT(计算!B9,"0.0")&" N"'),
        (11, "工作推力 N", "=计算!B6",
         "F = |Fg + Ff + Fa + F外|",
         '="("&TEXT(计算!B7,"0.0")&") + ("&TEXT(计算!B8,"0.0")&") + ("&TEXT(计算!B9,"0.0")&") + ("&TEXT(输入!B8,"0.0")&") = "&TEXT(计算!B45,"0.0")&" → |F|="&TEXT(计算!B6,"0.0")&" N"'),
        (12, "是否倒拖", "=计算!B10",
         "重力沿走向与运动相反时为倒拖",
         '=计算!B10&IF(计算!B10="是","　必须抱闸","　水平一般不倒拖，仍建议刹车")'),
        (14, "推荐丝杠", "=计算!B12",
         "目录第一档现货且转速、临界、轴径都过的",
         '=计算!B12&"　直径 "&TEXT(INDEX(目录丝杠!B2:B13,计算!B11),"0")&"　导程 "&TEXT(计算!B40,"0")&" mm"'),
        (15, "丝杠转速 r/min", "=计算!B13",
         "n = v(m/min)×1000 / 导程(mm)",
         '=TEXT(输入!B9,"0.00")&" × 1000 / "&TEXT(计算!B40,"0")&" = "&TEXT(计算!B13,"0.0")&" r/min"'),
        (16, "丝杠转矩 N·m", "=计算!B14",
         "T = F × P / (2π η)，P 为导程(m) 即 导程mm/1000",
         '="F="&TEXT(计算!B6,"0.0")&"，P="&TEXT(计算!B40,"0")&" mm，η="&TEXT(输入!B15,"0.00")&"，2πη="&TEXT(计算!B43,"0.000")&" → T="&TEXT(计算!B14,"0.000")&" N·m"'),
        (17, "丝杠机械功率 kW", "=计算!B36",
         "P = T n / 9550",
         '=TEXT(计算!B14,"0.000")&" × "&TEXT(计算!B13,"0.0")&" / 9550 = "&TEXT(计算!B36,"0.000")&" kW"'),
        (18, "轴承座", '=计算!B16&" + "&计算!B17',
         "按丝杠直径配 BK 固定 + BF 游动",
         '=计算!B16&" + "&计算!B17&"；BK内 "&计算!B18&"；BF内 "&计算!B19'),
        (19, "推荐伺服（丝杠）", "=计算!B30",
         "目录中运行 SF≥目标 且 启动 SF≥1.5 的最小一档",
         '=计算!B30&"　目标SF="&TEXT(输入!B13,"0.00")'),
        (20, "推荐步进（丝杠）", "=计算!B31",
         "按该转速掉矩后仍满足安全系数",
         '=计算!B31'),
        (22, "推荐齿条", '=TEXT(输入!B18,"0.0")&"模 × "&TEXT(输入!B19,"0")&"齿 "&输入!B16&"　i="&TEXT(输入!B20,"0")',
         "模数、齿数、减速比来自输入页",
         '="节圆半径 r="&TEXT(计算!B44,"0.0")&" mm（= m z / 2）"'),
        (23, "齿轮轴转速 r/min", "=计算!B39",
         "n0 = v×1000 / (π m z)",
         '=TEXT(输入!B9,"0.00")&"×1000 / (π×"&TEXT(输入!B18,"0.0")&"×"&TEXT(输入!B19,"0")&") = "&TEXT(计算!B39,"0.00")&" r/min"'),
        (24, "齿条电机转速 r/min", "=计算!B20",
         "n电机 = n0 × i",
         '=TEXT(计算!B39,"0.00")&" × "&TEXT(输入!B20,"0")&" = "&TEXT(计算!B20,"0.00")&" r/min"'),
        (25, "齿条电机转矩 N·m", "=计算!B21",
         "T电机 = F×(m z /2000) / η齿 / i / η减",
         '="F="&TEXT(计算!B6,"0.0")&"，mz/2000="&TEXT(输入!B18*输入!B19/2000,"0.000")&"，η齿="&TEXT(输入!B17,"0.00")&"，i="&TEXT(输入!B20,"0")&"，η减="&TEXT(IF(输入!B20<=1,1,输入!B21),"0.00")&" → T="&TEXT(计算!B21,"0.000")&" N·m"'),
        (26, "齿条机械功率 kW", "=计算!B37",
         "P = T电机 n电机 / 9550",
         '=TEXT(计算!B21,"0.000")&" × "&TEXT(计算!B20,"0.0")&" / 9550 = "&TEXT(计算!B37,"0.000")&" kW"'),
        (27, "推荐伺服（齿条）", "=计算!B32",
         "按减速后的 n、T 再选",
         '=计算!B32'),
        (28, "推荐步进（齿条）", "=计算!B33",
         "按减速后的 n、T 再选",
         '=计算!B33'),
    ]
    for i, h in enumerate(["项目", "结果", "公式", "代入（数字随输入变）"], 1):
        sty(wsR.cell(3, i, h), H, GREEN)
    for row, lab, val, formula, subst in rec:
        wsR.cell(row, 1, lab)
        wsR.cell(row, 2, val)
        wsR.cell(row, 3, formula)
        wsR.cell(row, 4, subst)
        sty(wsR.cell(row, 1), N, GRAY, False, True)
        sty(wsR.cell(row, 2), N, GREEN, False, True, "0.000")
        sty(wsR.cell(row, 3), Font(name="微软雅黑", size=10), GRAY, False, True)
        sty(wsR.cell(row, 4), N, GREEN, False, True)
        wsR.row_dimensions[row].height = 36
    wsR["A13"] = "↓ 丝杠方案"
    sty(wsR["A13"], H, BLUE)
    wsR.merge_cells("A13:D13")
    wsR["A21"] = "↓ 齿条方案"
    sty(wsR["A21"], H, BLUE)
    wsR.merge_cells("A21:D21")
    wsR["A30"] = "传动选「只要丝杠」时，齿条几行仍会算，只是不当采购依据。手选型号到「对照」看安全系数。改输入后本页数字一起变。"
    sty(wsR["A30"], Font(name="微软雅黑", size=10, color="666666"), wrap=True)
    wsR.merge_cells("A30:D30")
    wsR.row_dimensions[30].height = 28
    colw(wsR, [22, 22, 48, 62])
    wsR.freeze_panes = "A4"

    # ---- 计算过程（给会审看步骤）----
    wsK = wb.create_sheet("计算过程", 2)
    wsK["A1"] = "非标设计　计算过程（随输入自动变，只读）"
    sty(wsK["A1"], TITLE)
    wsK.merge_cells("A1:D1")
    wsK["A2"] = "先看本页步骤，再看「推荐」「对照」。公式与「计算」页同一套。"
    sty(wsK["A2"], N)
    wsK.merge_cells("A2:D2")
    for i, h in enumerate(["步骤", "公式", "代入与结果", "说明"], 1):
        sty(wsK.cell(4, i, h), H, BLUE)

    steps = [
        (5, "1 总质量", "m = 工件 + 机构",
         '=TEXT(输入!B6,"0.00")&" + "&TEXT(输入!B7,"0.00")&" = "&TEXT(计算!B2,"0.00")&" kg"',
         "架子请填实重，不要用估计的 640 kg"),
        (6, "2 速度", "v = 速度(m/min) ÷ 60",
         '=TEXT(输入!B9,"0.00")&" / 60 = "&TEXT(计算!B3,"0.000")&" m/s"',
         "后面加速度、功率都用 m/s"),
        (7, "3 加速度", "a = v ÷ 加速时间",
         '=TEXT(计算!B3,"0.000")&" / "&TEXT(输入!B11,"0.00")&" = "&TEXT(计算!B4,"0.000")&" m/s²"',
         "时间越短，加速力越大"),
        (8, "4 等效倾角", "水平0°；垂直上+90°；垂直下−90°；倾斜取填角（下坡为负）",
         '=输入!B4&" → "&TEXT(计算!B5,"0.0")&" °"',
         "倾斜必须在输入页填倾角"),
        (9, "5 沿程重力", "Fg = m g sinθ",
         '=TEXT(计算!B2,"0.00")&" × 9.81 × SIN("&TEXT(计算!B5,"0.0")&"°) = "&TEXT(计算!B2,"0.00")&" × 9.81 × "&TEXT(计算!B41,"0.000")&" = "&TEXT(计算!B7,"0.0")&" N"',
         "水平上坡为正、下坡为负"),
        (10, "6 摩擦力", "Ff = μ m g |cosθ|",
         '=TEXT(输入!B12,"0.00")&" × "&TEXT(计算!B2,"0.00")&" × 9.81 × |"&TEXT(计算!B42,"0.000")&"| = "&TEXT(计算!B8,"0.0")&" N"',
         "铜套 μ≈0.1，导轨 μ≈0.02"),
        (11, "7 加速力", "Fa = m a",
         '=TEXT(计算!B2,"0.00")&" × "&TEXT(计算!B4,"0.000")&" = "&TEXT(计算!B9,"0.0")&" N"',
         ""),
        (12, "8 工作推力", "F = |Fg + Ff + Fa + F外|",
         '="("&TEXT(计算!B7,"0.0")&") + ("&TEXT(计算!B8,"0.0")&") + ("&TEXT(计算!B9,"0.0")&") + ("&TEXT(输入!B8,"0.0")&") = "&TEXT(计算!B45,"0.0")&" → |F|="&TEXT(计算!B6,"0.0")&" N"',
         "电机按这个 F 来算转矩"),
        (13, "9 是否倒拖", "重力沿走向与运动相反时为倒拖",
         '=计算!B10',
         "倒拖必须抱闸"),
        (14, "10 丝杠型号", "目录中第一档「可用」现货",
         '=计算!B12',
         "优先 2510 / 4010"),
        (15, "11 丝杠转速", "n = v(m/min)×1000 / 导程(mm)",
         '=TEXT(输入!B9,"0.00")&" × 1000 / "&TEXT(计算!B40,"0")&" = "&TEXT(计算!B13,"0.0")&" r/min"',
         "导程 10 时，8 m/min → 800 r/min"),
        (16, "12 丝杠转矩", "T = F × P / (2π η)，P 为导程(m)=导程mm/1000",
         '="F="&TEXT(计算!B6,"0.0")&"，P="&TEXT(计算!B40,"0")&" mm，η="&TEXT(输入!B15,"0.00")&"，2πη="&TEXT(计算!B43,"0.000")&" → T="&TEXT(计算!B6,"0.0")&"×"&TEXT(计算!B40,"0")&"/1000/"&TEXT(计算!B43,"0.000")&" = "&TEXT(计算!B14,"0.000")&" N·m"',
         "效率默认 0.90"),
        (17, "13 丝杠功率", "P = T n / 9550",
         '=TEXT(计算!B14,"0.000")&" × "&TEXT(计算!B13,"0.0")&" / 9550 = "&TEXT(计算!B36,"0.000")&" kW"',
         "这是机械功率，电机按目录再留安全系数"),
        (18, "14 轴承座", "按丝杠直径配 BK 固定 + BF 游动",
         '=计算!B16&" + "&计算!B17&"；BK内 "&计算!B18&"；BF内 "&计算!B19',
         "BK 是一对角接触，BF 是一只深沟球"),
        (19, "15 齿轮轴转速", "n0 = v×1000 / (π m z)",
         '=TEXT(输入!B9,"0.00")&"×1000 / (π×"&TEXT(输入!B18,"0.0")&"×"&TEXT(输入!B19,"0")&") = "&TEXT(计算!B39,"0.00")&" r/min"',
         "减速前，齿轮转得慢"),
        (20, "16 齿条电机转速", "n电机 = n0 × i",
         '=TEXT(计算!B39,"0.00")&" × "&TEXT(输入!B20,"0")&" = "&TEXT(计算!B20,"0.00")&" r/min"',
         "i=1 为直连；默认 i=10"),
        (21, "17 齿条电机转矩", "T电机 = F×(m z /2000) / η齿 / i / η减",
         '="F="&TEXT(计算!B6,"0.0")&" × "&TEXT(输入!B18*输入!B19/2000,"0.000")&" / "&TEXT(输入!B17,"0.00")&" / "&TEXT(输入!B20,"0")&" / "&TEXT(IF(输入!B20<=1,1,输入!B21),"0.00")&" = "&TEXT(计算!B21,"0.000")&" N·m"',
         "直连 i=1 时 T 很大，目录会无满足项"),
        (22, "18 齿条功率", "P = T电机 n电机 / 9550",
         '=TEXT(计算!B21,"0.000")&" × "&TEXT(计算!B20,"0.0")&" / 9550 = "&TEXT(计算!B37,"0.000")&" kW"',
         "与丝杠功率应接近（差在效率）"),
        (23, "19 丝杠用伺服", "目录中运行 SF≥目标 且 启动 SF≥1.5 的最小一档",
         '=计算!B30',
         "目标安全系数见输入页"),
        (24, "20 丝杠用步进", "按该转速掉矩后仍满足安全系数",
         '=计算!B31',
         "转速高时 86 可能不够，改 110"),
        (25, "21 齿条用伺服", "按减速后的 n、T 再选",
         '=计算!B32',
         "仍无满足项就加大 i"),
        (26, "22 齿条用步进", "同上",
         '=计算!B33',
         ""),
    ]
    for row, a, b, c, d in steps:
        wsK.cell(row, 1, a)
        wsK.cell(row, 2, b)
        wsK.cell(row, 3, c)
        wsK.cell(row, 4, d)
        sty(wsK.cell(row, 1), H, GRAY, False, True)
        sty(wsK.cell(row, 2), N, GRAY, False, True)
        sty(wsK.cell(row, 3), N, GREEN, False, True)
        sty(wsK.cell(row, 4), Font(name="微软雅黑", size=10, color="555555"), GRAY, False, True)
        wsK.row_dimensions[row].height = 32
    colw(wsK, [18, 42, 55, 36])
    wsK.freeze_panes = "A5"

    # ---- 对照 ----
    wsD = wb.create_sheet("对照", 3)
    wsD["A1"] = "非标设计　丝杠方案 vs 齿轮齿条方案（同一载荷、同一速度）"
    sty(wsD["A1"], TITLE)
    wsD.merge_cells("A1:C1")
    headers = ["项目", "滚珠丝杠", "齿轮齿条"]
    for i, h in enumerate(headers, 1):
        sty(wsD.cell(3, i, h), H, BLUE)
    rows_d = [
        (4, "效率", "=输入!B15", "=输入!B17*IF(输入!B20<=1,1,输入!B21)"),
        (5, "减速比", "1（丝杠直连）", "=输入!B20"),
        (6, "电机转速 r/min", "=计算!B13", "=计算!B20"),
        (7, "需要转矩 N·m", "=计算!B14", "=计算!B21"),
        (8, "机械功率 kW", "=计算!B36", "=计算!B37"),
        (9, "推荐伺服", "=计算!B30", "=计算!B32"),
        (10, "推荐步进", "=计算!B31", "=计算!B33"),
        (11, "支承", '=计算!B16&" + "&计算!B17', "减速机+齿轮轴轴承（不是BK座）"),
        (12, "行程适应性", '=IF(输入!B10>4000,"长行程丝杠易垂","短中行程合适")', '=IF(输入!B10>4000,"长行程更合适","短行程也能用")'),
        (13, "效率谁高", "", '=IF(输入!B17*IF(输入!B20<=1,1,输入!B21)>输入!B15,"齿条链效率更高","丝杠效率更高或相当")'),
    ]
    for row, lab, a, b in rows_d:
        wsD.cell(row, 1, lab)
        wsD.cell(row, 2, a)
        wsD.cell(row, 3, b)
        sty(wsD.cell(row, 1), N, GRAY)
        sty(wsD.cell(row, 2), N)
        sty(wsD.cell(row, 3), N)
    wsD["A15"] = "手选后（与推荐可能不同）"
    sty(wsD["A15"], H)
    for i, h in enumerate(["项目", "手选丝杠方案", "手选齿条方案"], 1):
        sty(wsD.cell(16, i, h), H, YELLOW)
    hand_rows = [
        (17, "型号/参数", "=计算!B27", '=TEXT(计算!B23,"0.0")&"×"&TEXT(计算!B24,"0")&" "&输入!B16&" i="&TEXT(计算!B38,"0")'),
        (18, "转速", "=计算!B28", "=计算!B25"),
        (19, "转矩", "=计算!B29", "=计算!B26"),
        (20, "伺服", "=计算!B34", "=计算!B34"),
        (21, "步进", "=计算!B35", "=计算!B35"),
    ]
    for row, lab, a, b in hand_rows:
        wsD.cell(row, 1, lab)
        wsD.cell(row, 2, a)
        wsD.cell(row, 3, b)
        sty(wsD.cell(row, 1), N, GRAY)
        sty(wsD.cell(row, 2), N, YELLOW)
        sty(wsD.cell(row, 3), N, YELLOW)
    wsD["A23"] = "手选电机的安全系数见「目录电机」里对应行的运行SF；低于目标安全系数不要当正式方案。齿条直连（减速比=1）力矩很大，目录可能仍无满足项，请把减速比改为 8～15。"
    sty(wsD["A22"], Font(name="微软雅黑", size=10), wrap=True)
    wsD.merge_cells("A23:C23")
    colw(wsD, [22, 42, 42])

    # ---- 会审 ----
    wsE = wb.create_sheet("会审", 4)
    wsE["A1"] = "非标设计　工艺 / 机械 / 计算机会审"
    sty(wsE["A1"], TITLE)
    wsE["A3"] = "机械工程师"
    sty(wsE["A3"], H, BLUE)
    wsE["A4"] = '=CONCATENATE("工作推力 ",TEXT(计算!B6,"0")," N，总质量 ",TEXT(计算!B2,"0.0")," kg。丝杠推荐 ",计算!B12,"，座 ",计算!B16,"+",计算!B17,"。齿条模数",TEXT(输入!B18,"0.0")," 齿数",TEXT(输入!B19,"0")," 减速比 i=",TEXT(输入!B20,"0"),"。BK只配丝杠，不要装到齿条上。")'
    wsE["A5"] = '=IF(计算!B10="是","垂直或下坡有倒拖，必须抱闸，BF端不要锁死。","水平无倒拖，仍建议电机带刹车，停电不溜。")'
    wsE["A6"] = '=IF(输入!B10>4000,"行程超过4米，机械建议优先齿条，丝杠要中间托。","行程不长，丝杠齿条都能做，对位要求高用丝杠。")'
    wsE["A8"] = "工艺工程师"
    sty(wsE["A8"], H, BLUE)
    wsE["A9"] = '=IF(AND(输入!B9>=15,计算!B13>1200),"满速15 m/min且丝杠转速高：86步进当常用速度不稳，试车按点动再提速。","按当前速度可以试车：先对导向，再上传动，禁止硬掰同轴。")'
    wsE["A10"] = "成套买轴承座，不要拆了用两套深沟球顶替角接触。垂直、倾斜先确认停电刹车能抱住。"
    wsE["A11"] = '=IF(输入!B7<50,"机构质量填得过小，核对架子实重再下单。","架子质量已手填，按实重校核有效。")'
    wsE["A13"] = "计算机工程师"
    sty(wsE["A13"], H, BLUE)
    wsE["A14"] = "只改「输入」黄格和下拉。计算、目录、推荐不要手改公式。导出或另存时文件名带日期。"
    wsE["A15"] = '=CONCATENATE("建议文件名：选型_",输入!B4,"_",TEXT(输入!B9,"0"),"mmin_",计算!B12)'
    wsE["A16"] = "本表是唯一公式源。不要再做另一套网页算法。CAD 源文件不放进这张表。"
    for r in range(4, 17):
        if wsE.cell(r, 1).value and r not in (3, 8, 13):
            sty(wsE.cell(r, 1), N, wrap=True)
            wsE.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            wsE.row_dimensions[r].height = 48
    colw(wsE, [80, 20, 20])

    # ---- 说明 ----
    wsP = wb.create_sheet("说明", 0)
    wsP["A1"] = "非标设计"
    sty(wsP["A1"], TITLE)
    lines = [
        "非标直线传动选型（伺服 / 步进 / 丝杠 / 齿轮齿条）。与托辊铝座无关。只改「输入」黄格。",
        "1. 只改「输入」黄色格子和下拉。",
        "2. 安装方向可选：水平、垂直向上、垂直向下、倾斜上坡、倾斜下坡。倾斜必须填倾角。",
        "3. 工件质量和机构质量分开填。不要用估计的 640 kg 当架子重。",
        "4. 看「推荐」：结果右边就是公式和代入数字。完整 22 步在「计算过程」。采购没有该型号时，在输入页手选，到「对照」看安全系数。",
        "5. 丝杠和齿轮齿条用同一推力对比。齿条请填减速比（默认 10）；直连=1 时力矩大，推荐常为无满足项。BK/BF 只属于丝杠。",
        "6. 会审页是工艺、机械、计算机三句意见，随计算结果变。",
        "7. 用 Excel 或 WPS 打开。若公式显示为文字，检查单元格格式为「常规」并允许计算。",
        "8. 目录里 3210/3220 标现货=0，自动推荐不会选它们，手选仍可选（常无货）。",
        "公式：F = mg·sinθ + μmg·cosθ + ma + F外；丝杠 T=F·P/(2πη)，n=v×1000/P；齿条齿轮 n0=v×1000/(π m z)，电机 n=n0×i，T=F·(mz/2000)/η/i/η减。",
    ]
    for i, t in enumerate(lines, 3):
        wsP.cell(i, 1, t)
        sty(wsP.cell(i, 1), N, wrap=True)
        wsP.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)
        wsP.row_dimensions[i].height = 28
    colw(wsP, [80, 16, 16, 16])

    # 保护：输入黄格已 unlock
    for ws in (wsI,):
        ws.protection.sheet = False  # 允许改黄格；若开启保护需 unlock。先不强制保护以免 WPS 麻烦
    wsI.protection.sheet = False

    order = ["说明", "输入", "推荐", "计算过程", "对照", "会审", "计算", "目录丝杠", "目录电机"]
    wb._sheets = [wb[n] for n in order]

    # 条件格式：对照转速过高
    wsI.sheet_view.showGridLines = True
    wb.active = wsI
    wb.save(OUT)
    print("wrote", OUT)


if __name__ == "__main__":
    main()
